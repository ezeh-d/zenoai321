"""Table intelligence (#23) -- tables as STRUCTURED data, not flattened text.

Extract tables from spreadsheets (openpyxl), delimited files (csv), and PDFs
(PyMuPDF's find_tables where available) into {headers, rows} with provenance
(sheet or page), and convert them to CSV / JSON / Markdown / XLSX. This is what
"extract this table", "send that table to Excel", "turn it into CSV" run on.

Honest: if a format has no table extractor available it says so; it never
fabricates a table. Deterministic (no model).
"""

from __future__ import annotations

import csv
import io
import json
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

_MAX_ROWS = 5000
_MAX_COLS = 200


@dataclass
class Table:
    headers: list[str]
    rows: list[list[str]]
    source: str = ""
    location: str = ""          # "Sheet1" | "page 3" | ...
    meta: dict[str, Any] = field(default_factory=dict)

    @property
    def shape(self) -> tuple[int, int]:
        return (len(self.rows), len(self.headers))

    def as_dict(self) -> dict[str, Any]:
        return {"headers": self.headers, "rows": self.rows,
                "row_count": len(self.rows), "col_count": len(self.headers),
                "location": self.location, "source": self.source}


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _bound(rows: list[list[str]]) -> list[list[str]]:
    return [r[:_MAX_COLS] for r in rows[:_MAX_ROWS]]


# --- extraction -------------------------------------------------------------
def _from_xlsx(path: Path) -> list[Table]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    tables: list[Table] = []
    try:
        for ws in wb.worksheets:
            grid = [[_clean(c) for c in row]
                    for row in ws.iter_rows(values_only=True)]
            grid = [r for r in grid if any(cell for cell in r)]
            if not grid:
                continue
            headers = grid[0]
            tables.append(Table(headers=headers[:_MAX_COLS],
                                rows=_bound(grid[1:]), source=str(path),
                                location=ws.title,
                                meta={"sheet": ws.title}))
    finally:
        wb.close()
    return tables


def _from_delimited(path: Path, delim: str) -> list[Table]:
    text = path.read_text(encoding="utf-8", errors="replace")
    rows = [[_clean(c) for c in r]
            for r in csv.reader(io.StringIO(text), delimiter=delim)]
    rows = [r for r in rows if any(cell for cell in r)]
    if not rows:
        return []
    return [Table(headers=rows[0][:_MAX_COLS], rows=_bound(rows[1:]),
                  source=str(path), location=path.name)]


def _from_pdf(path: Path) -> list[Table] | None:
    """Uses PyMuPDF's table finder when present. Returns None (not []) when the
    capability isn't available, so the caller reports honestly."""
    try:
        import fitz
    except Exception:  # noqa: BLE001
        return None
    tables: list[Table] = []
    try:
        with fitz.open(path) as doc:
            for pno in range(doc.page_count):
                page = doc.load_page(pno)
                finder = getattr(page, "find_tables", None)
                if finder is None:
                    return None            # older PyMuPDF: no table finder
                for tbl in finder().tables:
                    data = tbl.extract()
                    data = [[_clean(c) for c in row] for row in data if any(row)]
                    if not data:
                        continue
                    tables.append(Table(headers=data[0][:_MAX_COLS],
                                        rows=_bound(data[1:]), source=str(path),
                                        location=f"page {pno + 1}"))
    except Exception:  # noqa: BLE001
        return tables or None
    return tables


def extract_tables(path: str | Path) -> dict[str, Any]:
    """All tables in a file, structured, with provenance. Honest on failure."""
    from reyes_agent.content import format_router as fr

    p = Path(os.path.abspath(os.path.expanduser(str(path))))
    if not p.exists() or not p.is_file():
        return {"ok": False, "error": f"'{p}' is not a file", "tables": []}
    info = fr.detect(p)
    try:
        if info.fmt in ("xlsx", "xlsm"):
            tables = _from_xlsx(p)
        elif info.fmt in ("csv",):
            tables = _from_delimited(p, ",")
        elif info.fmt in ("tsv",):
            tables = _from_delimited(p, "\t")
        elif info.fmt == "pdf":
            tables = _from_pdf(p)
            if tables is None:
                return {"ok": False, "format": info.fmt, "tables": [],
                        "error": "PDF table extraction needs a newer PyMuPDF "
                                 "(find_tables); not available here"}
        else:
            return {"ok": False, "format": info.fmt, "tables": [],
                    "error": f"no table extractor for '{info.fmt}'"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "format": info.fmt, "tables": [],
                "error": f"{type(exc).__name__}: {exc}"[:200]}

    return {"ok": bool(tables), "format": info.fmt, "count": len(tables),
            "tables": [t.as_dict() for t in tables],
            "error": "" if tables else "no tables found in this file"}


# --- conversion -------------------------------------------------------------
def to_csv(table: Table | dict) -> str:
    t = _coerce(table)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(t.headers)
    writer.writerows(t.rows)
    return buf.getvalue()


def to_json(table: Table | dict) -> str:
    t = _coerce(table)
    records = [dict(zip(t.headers, row + [""] * (len(t.headers) - len(row))))
               for row in t.rows]
    return json.dumps(records, ensure_ascii=False, indent=2)


def to_markdown(table: Table | dict) -> str:
    t = _coerce(table)
    if not t.headers:
        return ""
    lines = ["| " + " | ".join(t.headers) + " |",
             "| " + " | ".join("---" for _ in t.headers) + " |"]
    for row in t.rows:
        cells = list(row) + [""] * (len(t.headers) - len(row))
        lines.append("| " + " | ".join(c.replace("|", "\\|") for c in cells[:len(t.headers)]) + " |")
    return "\n".join(lines)


def save_table(table: Table | dict, dest: str | Path) -> dict[str, Any]:
    """Write a table to CSV / JSON / Markdown / XLSX by the destination's
    extension, then verify the write. Never claims success unverified (#18)."""
    t = _coerce(table)
    d = Path(os.path.abspath(os.path.expanduser(str(dest))))
    ext = d.suffix.lower()
    try:
        d.parent.mkdir(parents=True, exist_ok=True)
        if ext in (".csv", ".txt"):
            d.write_text(to_csv(t), encoding="utf-8")
        elif ext == ".json":
            d.write_text(to_json(t), encoding="utf-8")
        elif ext in (".md", ".markdown"):
            d.write_text(to_markdown(t), encoding="utf-8")
        elif ext in (".xlsx",):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(t.headers)
            for row in t.rows:
                ws.append(list(row))
            wb.save(d)
        else:
            return {"ok": False, "error": f"unsupported table destination '{ext}'"}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": f"{type(exc).__name__}: {exc}"[:200]}

    from reyes_agent.content.save import verify_write
    v = verify_write(d)
    return {"ok": v["ok"], "path": str(d), "format": ext.lstrip("."),
            "verified": v["checks"], "rows": len(t.rows), "cols": len(t.headers)}


def _coerce(table: Table | dict) -> Table:
    if isinstance(table, Table):
        return table
    return Table(headers=list(table.get("headers", [])),
                 rows=[list(r) for r in table.get("rows", [])],
                 source=table.get("source", ""), location=table.get("location", ""))
